# -*- coding: utf-8 -*-

"""
キャンペーン参加者 集計処理

【処理内容】

1. 「集計前」シートのB列が「取消済」の行を除外
2. A列（名前）＋C列（メール）をキーにしてD列（金額）を合計
3. 同じExcelブック内に「集計後」シートを作成
4. 合計金額の降順で並べ替え
5. 集計後の合計金額を基準に付与額を判定
6. D列の付与額ごとに色分け
7. F〜G列の付与額サマリを付与額降順
   （4万円→3万円→2万円→1万円）で出力
8. F:Gに付与額ごとの件数を表示
9. A, B, C, D, F, G列のみ自動調整
10. A〜D列、F〜G列に罫線設定
11. 実行前にバックアップを作成
"""

import re
import shutil
import unicodedata

from pathlib import Path
from datetime import datetime

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill,
    Border,
    Side,
    Font,
)

# ============================================================
# ユーザー設定
# ============================================================

BASE_DIR = Path(__file__).resolve().parent

INPUT_XLSX = BASE_DIR / "キャンペーン参加者.xlsx"
if not INPUT_XLSX.exists():
    raise FileNotFoundError(
        f"Excelファイルが見つかりません: {INPUT_XLSX}"
    )
INPUT_SHEET_NAME = "集計前"
OUTPUT_SHEET_NAME = "集計後"

# ============================================================
# 色設定
# ============================================================

COLOR_MAP = {
    "": "FFFFFF",
    "1万円": "C6EFCE",
    "2万円": "C9DAF8",
    "3万円": "FCE4D6",
    "4万円": "D9D9D9",
}

# ============================================================
# 金額変換
# ============================================================


def parse_money_to_yen(
    value,
    excel_row: int,
    warnings: list[str]
) -> int:

    if pd.isna(value):
        return 0

    if isinstance(value, (int, float)):
        return int(value)

    text = str(value).strip()

    if text == "":
        return 0

    text = re.sub(
        r"[,\s円¥￥$]",
        "",
        text
    )

    if text == "":
        return 0

    try:
        return int(float(text))

    except ValueError:

        warnings.append(
            f"{excel_row}行目: "
            f"'{value}' は数値化できないため0として処理しました"
        )

        return 0


# ============================================================
# 付与額判定
# ============================================================


def compute_tier_text(amount_yen: int) -> str:

    if amount_yen < 1_000_000:
        return ""

    if amount_yen < 2_000_000:
        return "1万円"

    if amount_yen < 3_000_000:
        return "2万円"

    if amount_yen < 4_000_000:
        return "3万円"

    return "4万円"


# ============================================================
# バックアップ
# ============================================================


def create_backup(file_path: Path) -> None:

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    backup_name = (
        f"{file_path.stem}_backup_{timestamp}"
        f"{file_path.suffix}"
    )

    backup_path = file_path.parent / backup_name

    shutil.copy2(file_path, backup_path)

    print(
        f"バックアップ作成: "
        f"{backup_path.name}"
    )

# ============================================================
# 集計
# ============================================================


def aggregate_data() -> tuple[pd.DataFrame, dict, list[str]]:

    warnings = []

    df = pd.read_excel(
        INPUT_XLSX,
        sheet_name=INPUT_SHEET_NAME,
        dtype=str
    )

    if df.shape[1] < 4:
        raise ValueError(
            "「集計前」シートにA～D列が必要です。"
        )

    colA, colB, colC, colD = df.columns[:4]

    # 「取消済」完全一致のみ除外

    df = df.loc[
        df[colB]
        .fillna("")
        .astype(str)
        .str.strip()
        != "取消済"
    ].copy()

    df["__金額_yen__"] = [
        parse_money_to_yen(
            value,
            idx + 2,
            warnings
        )
        for idx, value in enumerate(df[colD])
    ]

    grouped = (
        df.groupby(
            [colA, colC],
            dropna=False
        )["__金額_yen__"]
        .sum()
        .reset_index()
        .rename(
            columns={
                colA: "名前",
                colC: "メール",
                "__金額_yen__": "合計金額",
            }
        )
        .sort_values(
            by="合計金額",
            ascending=False
        )
        .reset_index(drop=True)
    )

    grouped["付与額"] = (
        grouped["合計金額"]
        .apply(compute_tier_text)
    )

    count_dict = (
        grouped["付与額"]
        .value_counts()
        .to_dict()
    )

    return grouped, count_dict, warnings

# ============================================================
# 列幅調整
# ============================================================


def get_display_width(text: str) -> int:

    width = 0

    for ch in text:

        if unicodedata.east_asian_width(ch) in ("F", "W", "A"):
            width += 2

        else:
            width += 1

    return width

# ============================================================
# シート作成
# ============================================================


def create_summary_sheet(
    grouped: pd.DataFrame,
    count_dict: dict
) -> None:

    wb = load_workbook(INPUT_XLSX)

    if OUTPUT_SHEET_NAME in wb.sheetnames:
        del wb[OUTPUT_SHEET_NAME]

    ws = wb.create_sheet(
        OUTPUT_SHEET_NAME,
        0
    )

    headers = [
        "名前",
        "メール",
        "合計金額",
        "付与額",
    ]

    for col_num, header in enumerate(
        headers,
        start=1
    ):
        ws.cell(
            row=1,
            column=col_num,
            value=header
        )

    for row_num, row_data in enumerate(
        grouped.itertuples(index=False),
        start=2
    ):

        ws.cell(row_num, 1, row_data.名前)
        ws.cell(row_num, 2, row_data.メール)
        ws.cell(row_num, 3, row_data.合計金額)
        ws.cell(row_num, 4, row_data.付与額)

    # 色分け

    for row_num in range(
        2,
        len(grouped) + 2
    ):

        tier = (
            ws.cell(
                row=row_num,
                column=4
            ).value
            or ""
        )

        color = COLOR_MAP.get(
            tier,
            "FFFFFF"
        )

        ws.cell(
            row=row_num,
            column=4
        ).fill = PatternFill(
            start_color=color,
            end_color=color,
            fill_type="solid"
        )

    # サマリ

    ws["F1"] = "付与額"
    ws["G1"] = "件数"

    tier_order = ["4万円", "3万円", "2万円", "1万円"]

    summary_sorted = [
        (tier, count_dict.get(tier, 0))
        for tier in tier_order
        if count_dict.get(tier, 0) > 0
    ]

    for i, (tier, cnt) in enumerate(
        summary_sorted,
        start=2
    ):

        ws[f"F{i}"] = tier
        ws[f"G{i}"] = f"{cnt}件"

        ws[f"F{i}"].fill = PatternFill(
            start_color=COLOR_MAP[tier],
            end_color=COLOR_MAP[tier],
            fill_type="solid"
        )

    # 列幅調整

    target_cols = ["A", "B", "C", "D", "F", "G"]

    for col_letter in target_cols:

        max_length = 0

        for cell in ws[col_letter]:
            if cell.value:

                for line in str(cell.value).split("\n"):

                    max_length = max(
                        max_length,
                        get_display_width(line)
                    )

        ws.column_dimensions[col_letter].width = (
        max_length + 2
        )

    # 罫線

    thin_side = Side(
        style="thin",
        color="000000"
    )

    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    last_row_ad = len(grouped) + 1
    last_row_fg = len(summary_sorted) + 1

    for row in range(
        1,
        last_row_ad + 1
    ):
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}{row}"].border = border

    for row in range(
        1,
        last_row_fg + 1
    ):
        for col in ["F", "G"]:
            ws[f"{col}{row}"].border = border

    # 太字

    bold_font = Font(bold=True)

    for col in ["A", "B", "C", "D", "F", "G"]:
        ws[f"{col}1"].font = bold_font

    wb.save(INPUT_XLSX)


# ============================================================
# メイン
# ============================================================


def main():

    if not INPUT_XLSX.exists():
        raise FileNotFoundError(
            f"Excelファイルが見つかりません: {INPUT_XLSX}"
        )

    create_backup(INPUT_XLSX)

    grouped, count_dict, warnings = aggregate_data()

    create_summary_sheet(
        grouped,
        count_dict
    )

    if warnings:

        print("\n=== 警告一覧 ===")

        for message in warnings:
            print(message)

    print("\n処理が完了しました。")


# ============================================================
# 実行
# ============================================================

if __name__ == "__main__":
    main()